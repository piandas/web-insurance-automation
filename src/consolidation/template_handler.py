"""
Manejador de plantillas Excel para diferentes fondos.

Este módulo se encarga de leer las plantillas Excel de los diferentes fondos
(EPM, FEPEP, etc.) y llenarlas con la información de cotizaciones.
"""

import os
import logging
from datetime import datetime
from typing import Dict, Any, Optional, List
from pathlib import Path
import openpyxl
import shutil
import unicodedata

from ..config.client_config import ClientConfig
from ..core.logger_factory import LoggerFactory


class TemplateHandler:
    """Manejador de plantillas Excel para diferentes fondos."""
    
    def __init__(self):
        self.logger = LoggerFactory.create_logger('template_handler')
        self.base_path = Path(__file__).parent.parent.parent
        self.templates_path = self.base_path / "Bases Consolidados"
        self.consolidados_path = self.base_path / "Consolidados"
        
        # Mapeo de fondos a archivos de plantilla (dinámico)
        self.template_files = self._discover_template_files()
        
        # Mapeo de fondos a aseguradoras que cotizan
        self.fondo_aseguradoras_map = {
            'FEPEP': ['SURA', 'ALLIANZ', 'BOLIVAR'],
            'CHEC': ['SURA', 'ALLIANZ', 'BOLIVAR'], 
            'EMVARIAS': ['SURA', 'BOLIVAR'],
            'EPM': ['SURA', 'ALLIANZ', 'BOLIVAR'],
            'CONFAMILIA': ['SURA', 'SOLIDARIA', 'BOLIVAR'],
            'FECORA': ['ALLIANZ', 'SOLIDARIA'],
            'FEMFUTURO': ['SBS', 'SOLIDARIA'],
            'FODELSA': ['SOLIDARIA'],
            'MANPOWER': ['SOLIDARIA', 'ALLIANZ', 'BOLIVAR']
        }
        
        # Crear directorio si no existe
        self.consolidados_path.mkdir(exist_ok=True)
    
    def _discover_template_files(self) -> Dict[str, str]:
        """Descubre automáticamente las plantillas disponibles."""
        templates = {}
        try:
            if self.templates_path.exists():
                # Buscar archivos que empiecen con "PANTILLA" o "PLANTILLA"
                for file_path in self.templates_path.glob("*ANTILLA*.xlsx"):
                    file_name = file_path.name
                    
                    # Filtrar archivos temporales de Excel (que empiecen con ~$)
                    if file_name.startswith("~$"):
                        continue
                    
                    # Extraer nombre del fondo del nombre del archivo
                    # Formato esperado: "PANTILLA FONDO N.xlsx" o "PLANTILLA FONDO U.xlsx"
                    if "PANTILLA" in file_name or "PLANTILLA" in file_name:
                        # Dividir por espacios y tomar la parte después de PANTILLA/PLANTILLA
                        parts = file_name.replace("PANTILLA", "").replace("PLANTILLA", "").strip()
                        fondo_name = parts.replace(".xlsx", "").strip()
                        if fondo_name:
                            templates[fondo_name] = file_name
                            self.logger.info(f"📋 Plantilla encontrada: {fondo_name} -> {file_name}")
        except Exception as e:
            self.logger.error(f"Error descubriendo plantillas: {e}")
        
        return templates
    
    def _normalize_text(self, text: str) -> str:
        """
        Normaliza texto removiendo tildes, acentos y convirtiendo a minúsculas.
        
        Args:
            text: Texto a normalizar
            
        Returns:
            str: Texto normalizado sin tildes en minúsculas
        """
        if not text:
            return ""
        
        # Convertir a string por si acaso
        text = str(text)
        
        # Remover tildes y acentos usando unicodedata
        text_normalized = unicodedata.normalize('NFD', text)
        text_without_accents = ''.join(c for c in text_normalized if unicodedata.category(c) != 'Mn')
        
        # Convertir a minúsculas
        return text_without_accents.lower()
    
    def _write_to_cell_safe(self, worksheet, row: int, column: int, value):
        """
        Escribe un valor en una celda de forma segura, manejando celdas fusionadas.
        
        Args:
            worksheet: Hoja de Excel
            row: Número de fila
            column: Número de columna
            value: Valor a escribir
        """
        try:
            cell = worksheet.cell(row=row, column=column)
            
            # Verificar si la celda está fusionada
            if hasattr(cell, 'coordinate'):
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Es una celda fusionada, escribir en la celda principal (top-left)
                        top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left_cell.value = value
                        self.logger.debug(f"📝 Escribiendo en celda fusionada {merged_range}: {value}")
                        return
            
            # No es una celda fusionada, escribir normalmente
            cell.value = value
            self.logger.debug(f"📝 Escribiendo en celda {cell.coordinate}: {value}")
            
        except Exception as e:
            self.logger.warning(f"❌ Error escribiendo en celda ({row}, {column}): {e}")
            # Intentar escribir directamente como fallback
            try:
                worksheet.cell(row=row, column=column).value = value
            except:
                self.logger.error(f"❌ Fallback también falló para celda ({row}, {column})")
    
    def _find_best_plan_match(self, target_plan: str, available_plans: Dict[str, str]) -> Optional[str]:
        """
        Encuentra la mejor coincidencia para un plan objetivo en la lista de planes disponibles.
        
        Args:
            target_plan: Nombre del plan objetivo (de la plantilla)
            available_plans: Diccionario de planes disponibles {nombre: valor}
            
        Returns:
            str: Valor del plan que mejor coincide, o None si no encuentra coincidencia
        """
        if not target_plan or not available_plans:
            return None
        
        # Normalizar el plan objetivo
        target_normalized = self._normalize_text(target_plan)
        
        # Primero buscar coincidencia exacta normalizada
        for plan_name, plan_value in available_plans.items():
            if self._normalize_text(plan_name) == target_normalized:
                self.logger.debug(f"🎯 Coincidencia exacta: '{target_plan}' = '{plan_name}'")
                return plan_value
        
        # Mapeos específicos más completos
        known_mappings = {
            # Mapeos de Allianz
            'autos esencial': ['autos esencial'],
            'autos esencial + total': ['autos esencial + totales'],
            'autos esencial + totales': ['autos esencial + totales', 'autos esencial + total'],
            'autos plus': ['autos plus'],
            'autos llave en mano': ['autos llave en mano'],
            
            # Mapeos de Sura actualizados con nueva nomenclatura
            'global franquicia': ['global franquicia', 'autos parcial', 'perdida parcial'],
            'autos global': ['autos global', 'plan autos global', 'perdida parcial 10-1 smlmv'],
            'autos clasico': ['autos clasico', 'plan autos clasico'],
            
            # Mapeos legacy para compatibilidad hacia atrás
            'plan autos global': ['autos global', 'global franquicia'],
            'plan autos clasico': ['autos clasico'],
            'perdida parcial 10-1 smlmv': ['autos global'],
            'perdida parcial': ['global franquicia', 'autos global'],
            'autos parcial': ['global franquicia'],
            
            # Variaciones adicionales
            'esencial': ['autos esencial'],
            'plus': ['autos plus'],
            'llave en mano': ['autos llave en mano'],
            'global': ['autos global', 'plan autos global', 'global franquicia'],
            'clasico': ['autos clasico', 'plan autos clasico']
        }
        
        # Buscar en mapeos conocidos
        for pattern, variations in known_mappings.items():
            if target_normalized == pattern:
                for variation in variations:
                    for plan_name, plan_value in available_plans.items():
                        if self._normalize_text(plan_name) == variation:
                            self.logger.debug(f"🎯 Mapeo conocido: '{target_plan}' -> '{plan_name}' vía '{pattern}'")
                            return plan_value
        
        # Buscar coincidencias por palabras clave (mejorado)
        target_keywords = set(target_normalized.split())
        best_match = None
        best_score = 0
        best_plan_name = None
        
        for plan_name, plan_value in available_plans.items():
            plan_keywords = set(self._normalize_text(plan_name).split())
            
            # Calcular intersección de palabras clave
            common_keywords = target_keywords.intersection(plan_keywords)
            score = len(common_keywords)
            
            # Palabras clave importantes tienen más peso
            important_keywords = {'esencial', 'plus', 'llave', 'mano', 'global', 'clasico', 'parcial', 'franquicia', 'total', 'totales'}
            important_matches = common_keywords.intersection(important_keywords)
            if important_matches:
                score += len(important_matches) * 2  # Doble peso para palabras importantes
            
            # Requiere al menos 1 palabra en común, pero con mejor scoring
            if score > 0 and score > best_score:
                best_match = plan_value
                best_score = score
                best_plan_name = plan_name
        
        if best_match:
            self.logger.debug(f"🎯 Mejor coincidencia por keywords: '{target_plan}' -> '{best_plan_name}' (score: {best_score})")
        else:
            self.logger.debug(f"❌ No se encontró coincidencia para: '{target_plan}' en {list(available_plans.keys())}")
        
        return best_match
    
    
    def get_available_fondos(self) -> List[str]:
        """Obtiene la lista de fondos disponibles basada en las plantillas existentes."""
        available_fondos = []
        processed_fondos = set()
        
        for fondo, template_file in self.template_files.items():
            template_path = self.templates_path / template_file
            if template_path.exists():
                # Extraer el fondo base (sin N o U)
                base_fondo = fondo
                if fondo.endswith(' N') or fondo.endswith(' U'):
                    base_fondo = fondo[:-2].strip()
                
                if base_fondo not in processed_fondos:
                    available_fondos.append(base_fondo)
                    processed_fondos.add(base_fondo)
            else:
                self.logger.warning(f"Plantilla no encontrada para {fondo}: {template_path}")
        
        return available_fondos
    
    def get_fondo_aseguradoras(self, fondo: str) -> List[str]:
        """
        Obtiene las aseguradoras que cotiza un fondo específico.
        
        Args:
            fondo: Nombre del fondo
            
        Returns:
            List[str]: Lista de aseguradoras que cotiza el fondo
        """
        return self.fondo_aseguradoras_map.get(fondo, ['SURA', 'ALLIANZ', 'BOLIVAR'])  # Default para fondos no mapeados
    
    def generate_filename(self) -> str:
        """Genera un nombre único para el archivo basado en la fecha actual."""
        today = datetime.now()
        base_name = f"Cotizacion{today.strftime('%d-%m-%y')}"
        extension = ".xlsx"
        
        # Verificar si el archivo ya existe y generar nombre único
        counter = 0
        filename = f"{base_name}{extension}"
        full_path = self.consolidados_path / filename
        
        while full_path.exists():
            counter += 1
            filename = f"{base_name}({counter}){extension}"
            full_path = self.consolidados_path / filename
            
        return filename
    
    def create_consolidado_from_template(self, fondo: str, sura_data: Dict[str, Any], 
                                       sura_plans: Dict[str, str], allianz_plans: Dict[str, str], 
                                       bolivar_solidaria_plans: Dict[str, str]) -> str:
        """
        Crea un consolidado usando la plantilla del fondo especificado.
        
        Args:
            fondo: Nombre del fondo (EPM, FEPEP, etc.)
            sura_data: Datos del cliente de Sura
            sura_plans: Planes de Sura
            allianz_plans: Planes de Allianz
            bolivar_solidaria_plans: Planes calculados de Bolívar y Solidaria
            
        Returns:
            str: Ruta del archivo Excel creado
        """
        self.logger.info(f"📊 Creando consolidado usando plantilla de {fondo}")
        
        # Determinar la plantilla correcta según el estado del vehículo
        vehicle_state = ClientConfig.VEHICLE_STATE.lower()
        template_key = None
        
        if vehicle_state == 'nuevo':
            template_key = f"{fondo} N"  # EPM N para nuevos
        elif vehicle_state == 'usado':
            template_key = f"{fondo} U"  # EPM U para usados
        else:
            # Fallback: buscar plantilla sin sufijo
            template_key = fondo
        
        # Verificar que la plantilla existe
        if template_key not in self.template_files:
            # Intentar fallback
            if fondo in self.template_files:
                template_key = fondo
                self.logger.warning(f"Plantilla específica '{template_key}' no encontrada, usando '{fondo}'")
            else:
                raise ValueError(f"Plantilla no encontrada para fondo: {fondo} (estado: {vehicle_state})")
        
        template_file = self.template_files[template_key]
        template_path = self.templates_path / template_file
        
        if not template_path.exists():
            raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")
        
        # Generar nombre del archivo de salida
        output_filename = self.generate_filename()
        output_path = self.consolidados_path / output_filename
        
        # Copiar plantilla como base
        shutil.copy2(template_path, output_path)
        self.logger.info(f"📋 Plantilla copiada a: {output_path}")
        
        # Abrir el archivo Excel
        try:
            workbook = openpyxl.load_workbook(output_path)
            worksheet = workbook.active  # Usar la primera hoja
            
            # Llenar datos según la estructura de la plantilla
            self._fill_template_data(worksheet, fondo, sura_data, sura_plans, allianz_plans, bolivar_solidaria_plans)
            
            # Guardar el archivo
            workbook.save(output_path)
            workbook.close()
            
            self.logger.info(f"✅ Consolidado creado exitosamente: {output_filename}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"❌ Error creando consolidado desde plantilla: {e}")
            # Si hay error, intentar limpiar el archivo parcial
            try:
                if output_path.exists():
                    output_path.unlink()
            except:
                pass
            raise
    
    def _fill_template_data(self, worksheet, fondo: str, sura_data: Dict[str, Any], 
                          sura_plans: Dict[str, str], allianz_plans: Dict[str, str], 
                          bolivar_solidaria_plans: Dict[str, str]):
        """
        Llena los datos en la plantilla Excel.
        
        Nueva estructura:
        - Datos del cliente: celdas merged de la C a la I
        - Valores cotizados: sistema de intersección basado en "VALOR A PAGAR (IVA INCLUIDO)"
        """
        try:
            self.logger.info(f"📝 Llenando datos en plantilla de {fondo}")
            
            # === PARTE 1: LLENAR DATOS DEL CLIENTE ===
            # Los datos van en las celdas merged de la C a la I (columna 3)
            self._fill_client_data(worksheet, sura_data)
            
            # === PARTE 2: LLENAR VALORES COTIZADOS ===
            # Usar sistema de intersección para valores cotizados
            self._fill_quoted_values(worksheet, fondo, sura_plans, allianz_plans, bolivar_solidaria_plans)
            
            # === PARTE 3: CONFIGURAR FECHAS Y CALCULAR DÍAS DE COBERTURA ===
            # Configurar fechas de vigencia y calcular días automáticamente
            self.logger.info("🔄 Iniciando configuración de fechas y días de cobertura...")
            self._setup_vigencia_dates_and_coverage(worksheet)
            
            # === PARTE 4: REEMPLAZAR CELDAS "VALOR ASEGURADO AUTO" ===
            # Buscar y reemplazar todas las celdas que contengan "VALOR ASEGURADO AUTO"
            self._replace_valor_asegurado_cells(worksheet)
            
            self.logger.info(f"✅ Datos llenados en plantilla de {fondo}")
            
        except Exception as e:
            self.logger.error(f"❌ Error llenando datos en plantilla: {e}")
            raise
    
    def _fill_client_data(self, worksheet, sura_data: Dict[str, Any]):
        """Llena los datos del cliente en las celdas merged (C-I)."""
        self.logger.info("📝 Llenando datos del cliente...")
        
        # 1. Fecha De Cotización
        fecha_row = self._find_cell_with_text(worksheet, 'fecha de cotización', 'fecha cotización')
        if fecha_row:
            fecha_actual = datetime.now().strftime('%d/%m/%Y')
            self._write_to_cell_safe(worksheet, fecha_row, 3, fecha_actual)  # Columna C
            self.logger.info(f"✅ Fecha de cotización: {fecha_actual}")
        
        # 2. Nombre Funcionario o asociado (dejar vacío por ahora)
        funcionario_row = self._find_cell_with_text(worksheet, 'nombre funcionario', 'funcionario o asociado')
        if funcionario_row:
            self._write_to_cell_safe(worksheet, funcionario_row, 3, "")
            self.logger.info(f"✅ Nombre funcionario (vacío)")
        
        # 3. C.C. Funcionario o asociado (dejar vacío por ahora)
        cc_funcionario_row = self._find_cell_with_text(worksheet, 'c.c. funcionario', 'cc funcionario')
        if cc_funcionario_row:
            self._write_to_cell_safe(worksheet, cc_funcionario_row, 3, "")
            self.logger.info(f"✅ CC funcionario (vacío)")
        
        # 4. Nombre Asegurado
        nombre_row = self._find_cell_with_text(worksheet, 'nombre asegurado')
        if nombre_row:
            nombre_completo = f"{sura_data.get('CLIENT_FIRST_NAME', '')} {sura_data.get('CLIENT_SECOND_NAME', '')} {sura_data.get('CLIENT_FIRST_LASTNAME', '')} {sura_data.get('CLIENT_SECOND_LASTNAME', '')}"
            self._write_to_cell_safe(worksheet, nombre_row, 3, nombre_completo.strip())
            self.logger.info(f"✅ Nombre asegurado: {nombre_completo.strip()}")
        
        # 5. C.C. Asegurado
        cc_asegurado_row = self._find_cell_with_text(worksheet, 'c.c. asegurado', 'cc asegurado')
        if cc_asegurado_row:
            cc_asegurado = sura_data.get('CLIENT_DOCUMENT_NUMBER', '')
            self._write_to_cell_safe(worksheet, cc_asegurado_row, 3, cc_asegurado)
            self.logger.info(f"✅ CC asegurado: {cc_asegurado}")
        
        # 6. Placa
        placa_row = self._find_cell_with_text(worksheet, 'placa')
        if placa_row:
            placa = ClientConfig.VEHICLE_PLATE
            self._write_to_cell_safe(worksheet, placa_row, 3, placa)
            self.logger.info(f"✅ Placa: {placa}")
        
        # 7. Modelo
        modelo_row = self._find_cell_with_text(worksheet, 'modelo')
        if modelo_row:
            modelo = ClientConfig.VEHICLE_MODEL_YEAR
            self._write_to_cell_safe(worksheet, modelo_row, 3, modelo)
            self.logger.info(f"✅ Modelo: {modelo}")
        
        # 8. Marca Y Tipo
        marca_row = self._find_cell_with_text(worksheet, 'marca y tipo', 'marca tipo')
        if marca_row:
            marca_tipo = f"{ClientConfig.VEHICLE_BRAND} - {ClientConfig.VEHICLE_REFERENCE}"
            self._write_to_cell_safe(worksheet, marca_row, 3, marca_tipo)
            self.logger.info(f"✅ Marca y tipo: {marca_tipo}")
        
        # 9. Clase (referencia escogida en fasecolda)
        clase_row = self._find_cell_with_text(worksheet, 'clase')
        if clase_row:
            clase = ClientConfig.VEHICLE_REFERENCE
            self._write_to_cell_safe(worksheet, clase_row, 3, clase)
            self.logger.info(f"✅ Clase: {clase}")
        
        # 10. Código Fasecolda (ambos códigos CF y CH)
        codigo_row = self._find_cell_with_text(worksheet, 'codigo fasecolda', 'fasecolda')
        if codigo_row:
            # Intentar obtener códigos extraídos automáticamente primero
            cf_code, ch_code = self._get_extracted_fasecolda_codes()
            
            # Si no hay códigos extraídos, usar los manuales como fallback
            if not cf_code:
                cf_code = ClientConfig.MANUAL_CF_CODE
                ch_code = ClientConfig.MANUAL_CH_CODE
                self.logger.info("📝 Usando códigos FASECOLDA manuales como fallback")
                
                # Verificar que los códigos manuales no estén vacíos
                if not cf_code or not ch_code:
                    self.logger.warning("⚠️ Códigos FASECOLDA manuales están vacíos también")
                    cf_code = cf_code or "N/A"
                    ch_code = ch_code or "N/A"
            else:
                self.logger.info("📝 Usando códigos FASECOLDA extraídos automáticamente")
            
            # Formato final
            codigo_completo = f"CF: {cf_code} / CH: {ch_code}"
            self._write_to_cell_safe(worksheet, codigo_row, 3, codigo_completo)
            self.logger.info(f"✅ Código Fasecolda: {codigo_completo}")
        
        # 11. Ciudad de circulación
        ciudad_row = self._find_cell_with_text(worksheet, 'ciudad de circulación', 'ciudad circulación')
        if ciudad_row:
            ciudad = ClientConfig.CLIENT_CITY
            self._write_to_cell_safe(worksheet, ciudad_row, 3, ciudad)
            self.logger.info(f"✅ Ciudad de circulación: {ciudad}")
        
        # 12. Valor asegurado
        valor_row = self._find_cell_with_text(worksheet, 'valor asegurado')
        if valor_row:
            valor_asegurado = self._get_valor_asegurado()
            if valor_asegurado:
                valor_formateado = self._format_currency(valor_asegurado)
                self._write_to_cell_safe(worksheet, valor_row, 3, valor_formateado)
                self.logger.info(f"✅ Valor asegurado: {valor_formateado}")
        
        # 13. Valor accesorios (dejar vacío por ahora)
        accesorios_row = self._find_cell_with_text(worksheet, 'valor accesorios', 'accesorios')
        if accesorios_row:
            self._write_to_cell_safe(worksheet, accesorios_row, 3, "")
            self.logger.info(f"✅ Valor accesorios (vacío)")
        
        # 14. Valor asegurado total
        valor_total_row = self._find_cell_with_text(worksheet, 'valor asegurado total', 'total asegurado')
        if valor_total_row:
            valor_asegurado = self._get_valor_asegurado()
            if valor_asegurado:
                valor_formateado = self._format_currency(valor_asegurado)
                self._write_to_cell_safe(worksheet, valor_total_row, 3, valor_formateado)
                self.logger.info(f"✅ Valor asegurado total: {valor_formateado}")
    
    def _fill_quoted_values(self, worksheet, fondo: str, sura_plans: Dict[str, str], 
                          allianz_plans: Dict[str, str], bolivar_solidaria_plans: Dict[str, str]):
        """
        Llena los valores cotizados usando el sistema de intersección.
        
        Busca la fila "VALOR A PAGAR (IVA INCLUIDO)" y las columnas de las aseguradoras
        para crear intersecciones donde colocar los valores.
        Solo llena aseguradoras que el fondo tiene permitidas.
        """
        self.logger.info(f"💰 Llenando valores cotizados para fondo {fondo} usando sistema de intersección...")
        
        # Obtener aseguradoras permitidas para el fondo
        aseguradoras_permitidas = self.get_fondo_aseguradoras(fondo)
        self.logger.info(f"🔒 Aseguradoras permitidas para {fondo}: {aseguradoras_permitidas}")
        
        # 1. Buscar la fila "VALOR A PAGAR (IVA INCLUIDO)"
        valor_pagar_row = self._find_cell_with_text_in_any_column(worksheet, 'valor a pagar', 'iva incluido')
        if not valor_pagar_row:
            self.logger.warning("❌ No se encontró la fila 'VALOR A PAGAR (IVA INCLUIDO)'")
            return
        
        self.logger.info(f"✅ Fila 'VALOR A PAGAR' encontrada: {valor_pagar_row}")
        
        # 2. Determinar mapeo de planes - SIEMPRE usar los 3 planes de Sura
        # Nueva nomenclatura: Global Franquicia, Autos Global, Autos Clásico
        plan_mapping = {
            'sura': ['Global Franquicia', 'Autos Global', 'Autos Clásico'],
            'allianz': ['Autos Esencial', 'Autos Esencial + Total', 'Autos Plus', 'Autos llave en mano']
        }
        
        # Mapeo de planes de Sura con nueva nomenclatura
        sura_plan_map = {
            'Global Franquicia': sura_plans.get('Global Franquicia', 'No encontrado'),
            'Autos Global': sura_plans.get('Autos Global', 'No encontrado'),
            'Autos Clásico': sura_plans.get('Autos Clásico', 'No encontrado')
        }
        
        # 3. Buscar columnas de SURA y llenar valores (si está permitida)
        if 'SURA' in aseguradoras_permitidas:
            self._fill_sura_values(worksheet, valor_pagar_row, sura_plan_map, plan_mapping['sura'])
        else:
            self.logger.info(f"⚠️ SURA omitida para fondo {fondo} (no permitida)")
        
        # 4. Buscar columnas de ALLIANZ y llenar valores (si está permitida)
        if 'ALLIANZ' in aseguradoras_permitidas:
            self._fill_allianz_values(worksheet, valor_pagar_row, allianz_plans, plan_mapping['allianz'])
        else:
            self.logger.info(f"⚠️ ALLIANZ omitida para fondo {fondo} (no permitida)")
        
        # 5. Llenar Bolívar y Solidaria si están permitidas
        self._fill_other_values(worksheet, valor_pagar_row, bolivar_solidaria_plans, aseguradoras_permitidas, fondo)
        
        # Nota: Los valores anualizados se calculan en _setup_vigencia_dates_and_coverage()
    
    def _fill_sura_values(self, worksheet, valor_pagar_row: int, sura_plan_map: Dict[str, str], 
                         expected_plans: List[str]):
        """Llena los valores de Sura en las columnas correspondientes."""
        self.logger.info("📊 Llenando valores de SURA...")
        
        # Buscar columnas de SURA
        sura_columns = self._find_company_columns(worksheet, 'sura')
        if not sura_columns:
            self.logger.warning("❌ No se encontraron columnas de SURA")
            return
        
        self.logger.info(f"✅ Columnas de SURA encontradas: {sura_columns}")
        
        # Mapear planes a columnas leyendo encabezados reales
        for column in sura_columns:
            # Buscar el encabezado de esta columna
            column_header = self._get_column_header(worksheet, column)
            
            if column_header:
                # Buscar la mejor coincidencia en los planes de Sura
                matched_value = self._find_best_plan_match(column_header, sura_plan_map)
                
                if matched_value and matched_value != 'No encontrado':
                    formatted_value = self._format_currency_from_string(matched_value)
                    self._write_to_cell_safe(worksheet, valor_pagar_row, column, formatted_value)
                    self.logger.info(f"✅ SURA '{column_header}' -> {formatted_value} en columna {column}")
                else:
                    self.logger.warning(f"⚠️ SURA '{column_header}': No se encontró coincidencia en columna {column}")
            else:
                self.logger.warning(f"⚠️ No se pudo leer encabezado de columna SURA {column}")
        
        # Buscar adicicionalmente por nombre específico "Global Franquicia"
        global_franquicia_col = self._find_column_by_header(worksheet, 'global franquicia')
        if global_franquicia_col:
            parcial_value = sura_plan_map.get('Autos Parcial', 'No encontrado')
            if parcial_value != 'No encontrado':
                formatted_value = self._format_currency_from_string(parcial_value)
                self._write_to_cell_safe(worksheet, valor_pagar_row, global_franquicia_col, formatted_value)
                self.logger.info(f"✅ SURA Global Franquicia (sinónimo): {formatted_value} en columna {global_franquicia_col}")
    
    def _fill_allianz_values(self, worksheet, valor_pagar_row: int, allianz_plans: Dict[str, str], 
                           expected_plans: List[str]):
        """Llena los valores de Allianz en las columnas correspondientes."""
        self.logger.info("📊 Llenando valores de ALLIANZ...")
        
        # Buscar columnas de ALLIANZ
        allianz_columns = self._find_company_columns(worksheet, 'allianz')
        if not allianz_columns:
            self.logger.warning("❌ No se encontraron columnas de ALLIANZ")
            return
        
        self.logger.info(f"✅ Columnas de ALLIANZ encontradas: {allianz_columns}")
        
        # Mapear planes a columnas leyendo encabezados reales
        for column in allianz_columns:
            # Buscar el encabezado de esta columna
            column_header = self._get_column_header(worksheet, column)
            
            if column_header:
                # Buscar la mejor coincidencia en los planes de Allianz
                matched_value = self._find_best_plan_match(column_header, allianz_plans)
                
                if matched_value and matched_value != 'No encontrado':
                    formatted_value = self._format_currency_from_string(matched_value)
                    self._write_to_cell_safe(worksheet, valor_pagar_row, column, formatted_value)
                    self.logger.info(f"✅ ALLIANZ '{column_header}' -> {formatted_value} en columna {column}")
                else:
                    self.logger.warning(f"⚠️ ALLIANZ '{column_header}': No se encontró coincidencia en columna {column}")
            else:
                self.logger.warning(f"⚠️ No se pudo leer encabezado de columna ALLIANZ {column}")
    
    def _fill_other_values(self, worksheet, valor_pagar_row: int, bolivar_solidaria_plans: Dict[str, str], 
                         aseguradoras_permitidas: List[str], fondo: str):
        """Llena valores de otras aseguradoras si están permitidas para el fondo."""
        self.logger.info(f"📊 Intentando llenar valores de otras aseguradoras para fondo {fondo}...")
        
        # Buscar columnas de BOLIVAR (solo si está permitida)
        if 'BOLIVAR' in aseguradoras_permitidas:
            bolivar_columns = self._find_company_columns(worksheet, 'bolivar')
            if bolivar_columns:
                bolivar_value = bolivar_solidaria_plans.get('Bolívar', 'No calculado')
                if bolivar_value != 'No calculado':
                    formatted_value = self._format_currency_from_string(bolivar_value)
                    self._write_to_cell_safe(worksheet, valor_pagar_row, bolivar_columns[0], formatted_value)
                    self.logger.info(f"✅ BOLIVAR: {formatted_value}")
        else:
            self.logger.info(f"⚠️ BOLIVAR omitida para fondo {fondo} (no permitida)")
        
        # Buscar columnas de SOLIDARIA (solo si está permitida)
        if 'SOLIDARIA' in aseguradoras_permitidas:
            solidaria_columns = self._find_company_columns(worksheet, 'solidaria')
            if solidaria_columns:
                solidaria_value = bolivar_solidaria_plans.get('Solidaria', 'No calculado')
                if solidaria_value != 'No calculado':
                    formatted_value = self._format_currency_from_string(solidaria_value)
                    self._write_to_cell_safe(worksheet, valor_pagar_row, solidaria_columns[0], formatted_value)
                    self.logger.info(f"✅ SOLIDARIA: {formatted_value}")
        else:
            self.logger.info(f"⚠️ SOLIDARIA omitida para fondo {fondo} (no permitida)")
        
        # Buscar columnas de SBS (solo si está permitida)
        if 'SBS' in aseguradoras_permitidas:
            sbs_columns = self._find_company_columns(worksheet, 'sbs')
            if sbs_columns:
                sbs_value = bolivar_solidaria_plans.get('SBS', 'No calculado')
                if sbs_value != 'No calculado':
                    formatted_value = self._format_currency_from_string(sbs_value)
                    self._write_to_cell_safe(worksheet, valor_pagar_row, sbs_columns[0], formatted_value)
                    self.logger.info(f"✅ SBS: {formatted_value}")
        else:
            self.logger.info(f"⚠️ SBS omitida para fondo {fondo} (no permitida)")
    
    def _fill_annualized_values(self, worksheet, valor_pagar_row: int):
        """
        Calcula y llena los valores anualizados usando la fórmula:
        (Valor a pagar) / (Días de cobertura) * 365
        """
        self.logger.info("📊 Calculando valores anualizados...")
        
        # 1. Buscar la fila "Días de Cobertura"
        dias_cobertura_row = self._find_cell_with_text_in_any_column(worksheet, 'días de cobertura', 'dias de cobertura')
        if not dias_cobertura_row:
            self.logger.warning("❌ No se encontró la fila 'Días de Cobertura'")
            return
        
        # 2. Buscar la fila "PRIMA ANUAL IVA INCLUIDO"
        prima_anual_row = self._find_cell_with_text_in_any_column(worksheet, 'prima anual', 'iva incluido')
        if not prima_anual_row:
            self.logger.warning("❌ No se encontró la fila 'PRIMA ANUAL IVA INCLUIDO'")
            return
        
        self.logger.info(f"✅ Fila 'Días de Cobertura': {dias_cobertura_row}")
        self.logger.info(f"✅ Fila 'PRIMA ANUAL IVA INCLUIDO': {prima_anual_row}")
        
        # 3. Recorrer las columnas y calcular valores anualizados
        for col in range(2, 20):  # Columnas B hasta S
            try:
                # Obtener valor a pagar de la fila correspondiente
                valor_pagar_cell = worksheet.cell(row=valor_pagar_row, column=col)
                valor_pagar = valor_pagar_cell.value
                
                # Obtener días de cobertura
                dias_cobertura_cell = worksheet.cell(row=dias_cobertura_row, column=col)
                dias_cobertura = dias_cobertura_cell.value
                
                # Solo calcular si ambos valores existen y son válidos
                if valor_pagar and dias_cobertura and str(valor_pagar).strip() != "":
                    # Limpiar valor a pagar (quitar $ y puntos)
                    valor_numerico = self._extract_numeric_value(str(valor_pagar))
                    dias_numerico = self._extract_numeric_value(str(dias_cobertura))
                    
                    if valor_numerico > 0 and dias_numerico > 0:
                        # Calcular valor anualizado: (Valor a pagar) / (Días de cobertura) * 365
                        valor_anualizado = (valor_numerico / dias_numerico) * 365
                        
                        # Formatear como moneda
                        valor_formateado = f"${valor_anualizado:,.0f}".replace(",", ".")
                        
                        # Colocar en la fila PRIMA ANUAL
                        self._write_to_cell_safe(worksheet, prima_anual_row, col, valor_formateado)
                        
                        self.logger.info(f"✅ Columna {col}: ${valor_numerico:,.0f} ÷ {dias_numerico} × 365 = {valor_formateado}")
                
            except Exception as e:
                self.logger.warning(f"⚠️ Error calculando valor anualizado en columna {col}: {e}")
                continue
    
    def _extract_numeric_value(self, value_str: str) -> float:
        """Extrae el valor numérico de una string, removiendo caracteres no numéricos."""
        try:
            if not value_str or not value_str.strip():
                return 0.0
            
            # Limpiar valor - quitar símbolos y mantener solo números
            clean_value = str(value_str).replace("$", "").replace(".", "").replace(",", "").replace(" ", "").strip()
            
            # Quitar cualquier carácter no numérico
            numeric_only = ''.join(c for c in clean_value if c.isdigit())
            
            self.logger.debug(f"🔢 Extrayendo: '{value_str}' → '{clean_value}' → '{numeric_only}'")
            
            # Convertir a float
            if numeric_only and numeric_only.isdigit():
                result = float(numeric_only)
                self.logger.debug(f"✅ Valor numérico extraído: {result}")
                return result
            else:
                self.logger.warning(f"⚠️ No se pudo extraer valor numérico de: '{value_str}'")
                return 0.0
                
        except Exception as e:
            self.logger.error(f"Error extrayendo valor numérico de '{value_str}': {e}")
            return 0.0
    
    def _find_cell_with_text_in_any_column(self, worksheet, *search_terms) -> Optional[int]:
        """
        Busca una celda que contenga alguno de los términos especificados en cualquier columna.
        Usa normalización de texto para ignorar tildes y acentos.
        
        Returns:
            int: Número de fila si se encuentra, None si no se encuentra
        """
        try:
            # Normalizar términos de búsqueda
            normalized_terms = [self._normalize_text(term) for term in search_terms]
            
            self.logger.debug(f"🔍 Buscando términos: {search_terms}")
            
            for row in range(1, 100):  # Buscar en las primeras 100 filas
                for col in range(1, 20):  # Buscar en las primeras 20 columnas
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_text_normalized = self._normalize_text(str(cell_value))
                        # Verificar si CUALQUIER término está en el texto de la celda
                        for normalized_term in normalized_terms:
                            if normalized_term in cell_text_normalized:
                                self.logger.debug(f"✅ Término '{normalized_term}' encontrado en fila {row}, columna {col}: '{cell_value}'")
                                return row
            
            self.logger.debug(f"❌ No se encontraron términos: {search_terms}")
            return None
        except Exception as e:
            self.logger.error(f"Error buscando texto en celdas: {e}")
            return None
    
    def _find_company_columns(self, worksheet, company_name: str) -> List[int]:
        """
        Busca las columnas de una aseguradora específica.
        Usa normalización de texto para ignorar tildes y acentos.
        
        Returns:
            List[int]: Lista de números de columna donde aparece la aseguradora
        """
        columns = []
        try:
            company_name_normalized = self._normalize_text(company_name)
            
            for row in range(1, 50):  # Buscar en las primeras 50 filas
                for col in range(1, 20):  # Buscar en las primeras 20 columnas
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_text_normalized = self._normalize_text(str(cell_value))
                        if company_name_normalized in cell_text_normalized:
                            if col not in columns:
                                columns.append(col)
            
            return sorted(columns)
        except Exception:
            return []

    def _find_column_by_header(self, worksheet, header_text: str) -> Optional[int]:
        """
        Busca una columna específica por su texto de encabezado.
        Usa normalización de texto para ignorar tildes y acentos.
        
        Args:
            worksheet: Hoja de Excel
            header_text: Texto del encabezado a buscar
            
        Returns:
            int: Número de columna si se encuentra, None si no se encuentra
        """
        try:
            header_normalized = self._normalize_text(header_text)
            
            # Buscar en las primeras filas (donde suelen estar los encabezados)
            for row in range(1, 20):  # Buscar en las primeras 20 filas
                for col in range(1, 30):  # Buscar en las primeras 30 columnas
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_text_normalized = self._normalize_text(str(cell_value))
                        if header_normalized in cell_text_normalized:
                            self.logger.debug(f"✅ Encabezado '{header_text}' encontrado en columna {col}")
                            return col
            
            self.logger.debug(f"❌ Encabezado '{header_text}' no encontrado")
            return None
        except Exception as e:
            self.logger.error(f"Error buscando encabezado '{header_text}': {e}")
            return None
    
    def _format_currency_from_string(self, value: str) -> str:
        """Formatea un valor string como moneda colombiana."""
        try:
            if not value or not value.strip() or value == 'No encontrado':
                return value
            
            # Si el valor ya viene con formato colombiano (punto para miles, coma para decimales)
            # Como: "337.753,02" -> solo quitar decimales y agregar $
            if ',' in value:
                # Formato colombiano: separar parte entera de decimales
                parts = value.split(',')
                integer_part = parts[0]  # "337.753"
                # Ignorar la parte decimal como pediste
                return f"${integer_part}"
            
            # Si no tiene coma, asumir que es solo la parte entera
            # Limpiar caracteres no numéricos excepto puntos (separadores de miles)
            clean_value = ''.join(c for c in value if c.isdigit() or c == '.')
            if not clean_value:
                return value
            
            # Si ya tiene formato con puntos como separadores de miles, mantenerlo
            if '.' in clean_value and len(clean_value.split('.')[-1]) == 3:
                # Ya está formateado correctamente
                return f"${clean_value}"
            
            # Si es solo números, formatear como moneda
            if clean_value.isdigit():
                num = int(clean_value)
                formatted = f"{num:,}".replace(",", ".")
                return f"${formatted}"
            
            # Fallback: devolver como está
            return f"${clean_value}"
            
        except Exception:
            return value
    
    def _find_cell_with_text(self, worksheet, *search_terms) -> Optional[int]:
        """
        Busca una celda que contenga alguno de los términos especificados en la columna A.
        Usa normalización de texto para ignorar tildes y acentos.
        
        Returns:
            int: Número de fila si se encuentra, None si no se encuentra
        """
        try:
            # Normalizar términos de búsqueda
            normalized_terms = [self._normalize_text(term) for term in search_terms]
            
            for row in range(1, 50):  # Buscar en las primeras 50 filas
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value:
                    cell_text_normalized = self._normalize_text(str(cell_value))
                    for normalized_term in normalized_terms:
                        if normalized_term in cell_text_normalized:
                            return row
            return None
        except Exception:
            return None
    
    def _get_valor_asegurado(self) -> Optional[str]:
        """Obtiene el valor asegurado desde la configuración."""
        try:
            # Usar directamente VEHICLE_INSURED_VALUE para evitar _load_gui_overrides()
            # que podría sobrescribir el valor manual ingresado tanto para nuevos como usados
            valor = ClientConfig.VEHICLE_INSURED_VALUE
            vehicle_state = ClientConfig.VEHICLE_STATE.lower()
            self.logger.info(f"💰 Usando valor asegurado para vehículo {vehicle_state}: {valor}")
            return valor
            
        except Exception as e:
            self.logger.error(f"Error obteniendo valor asegurado: {e}")
            return None
    
    def _format_currency(self, value: str) -> str:
        """Formatea un valor como moneda colombiana."""
        try:
            if not value or not value.strip():
                return ""
            
            # Limpiar el valor (quitar caracteres no numéricos)
            clean_value = ''.join(c for c in value if c.isdigit())
            if not clean_value:
                return ""
            
            # Convertir a número y formatear
            num = int(clean_value)
            formatted = f"{num:,}".replace(",", ".")
            return f"${formatted}"
            
        except Exception:
            return value

    def _replace_valor_asegurado_cells(self, worksheet):
        """
        Busca y reemplaza todas las celdas que contengan exactamente 'VALOR ASEGURADO AUTO'
        por el valor asegurado real del vehículo.
        """
        try:
            valor_asegurado = self._get_valor_asegurado()
            if not valor_asegurado:
                self.logger.warning("No se pudo obtener el valor asegurado para reemplazar")
                return
            
            valor_formateado = self._format_currency(valor_asegurado)
            reemplazos_realizados = 0
            
            self.logger.info(f"🔍 Buscando celdas con 'VALOR ASEGURADO AUTO' para reemplazar con: {valor_formateado}")
            
            # Buscar en un rango amplio de celdas
            for row in range(1, 200):  # Buscar en las primeras 200 filas
                for col in range(1, 30):  # Buscar en las primeras 30 columnas
                    try:
                        cell = worksheet.cell(row=row, column=col)
                        if cell.value:
                            cell_text = str(cell.value).strip().upper()
                            # Buscar texto que contenga exactamente "VALOR ASEGURADO AUTO"
                            if "VALOR ASEGURADO AUTO" in cell_text:
                                # Reemplazar completamente el contenido de la celda
                                cell.value = valor_formateado
                                reemplazos_realizados += 1
                                self.logger.info(f"✅ Reemplazado en celda {cell.coordinate}: '{cell_text}' → '{valor_formateado}'")
                    except Exception as e:
                        # Continuar si hay error en una celda específica
                        continue
            
            if reemplazos_realizados > 0:
                self.logger.info(f"✅ Total de reemplazos realizados: {reemplazos_realizados}")
            else:
                self.logger.warning("⚠️ No se encontraron celdas con 'VALOR ASEGURADO AUTO' para reemplazar")
                
        except Exception as e:
                self.logger.error(f"Error al reemplazar celdas de VALOR ASEGURADO AUTO: {e}")

    def _get_column_header(self, worksheet, column: int) -> Optional[str]:
        """
        Obtiene el encabezado de una columna específica.
        Busca en las primeras filas el texto que mejor describe el plan.
        """
        try:
            # Buscar en las primeras 30 filas
            potential_headers = []
            
            for row in range(1, 31):
                cell_value = worksheet.cell(row=row, column=column).value
                if cell_value and str(cell_value).strip():
                    header_text = str(cell_value).strip()
                    
                    # Filtrar encabezados que parecen nombres de planes
                    if len(header_text) > 3 and not header_text.isdigit():
                        # Ignorar fechas (formato DD/MM/YYYY)
                        if '/' in header_text and len(header_text.split('/')) == 3:
                            continue
                        
                        # Ignorar encabezados genéricos
                        generic_headers = ['allianz', 'sura', 'bolivar', 'solidaria', 'coberturas', 'aseguradora', 'valor', 'prima', 'iva', 'incluido']
                        if self._normalize_text(header_text) in [self._normalize_text(g) for g in generic_headers]:
                            continue
                        
                        # Buscar términos que indican nombres de planes
                        plan_keywords = ['autos', 'plus', 'esencial', 'llave', 'mano', 'global', 'clasico', 'parcial', 'franquicia', 'total', 'totales']
                        if any(keyword in self._normalize_text(header_text) for keyword in plan_keywords):
                            potential_headers.append((row, header_text))
            
            # Si encontramos múltiples candidatos, usar el que esté más arriba en la plantilla
            if potential_headers:
                # Ordenar por fila (más arriba primero) 
                potential_headers.sort(key=lambda x: x[0])
                best_header = potential_headers[0][1]
                self.logger.debug(f"📋 Encabezado columna {column}: '{best_header}' (fila {potential_headers[0][0]})")
                return best_header
            
            # Si no encontramos nada específico, buscar cualquier texto no genérico
            for row in range(1, 31):
                cell_value = worksheet.cell(row=row, column=column).value
                if cell_value and str(cell_value).strip():
                    header_text = str(cell_value).strip()
                    if len(header_text) > 3 and not header_text.isdigit() and '/' not in header_text:
                        generic_headers = ['allianz', 'sura', 'bolivar', 'solidaria', 'coberturas', 'aseguradora']
                        if self._normalize_text(header_text) not in [self._normalize_text(g) for g in generic_headers]:
                            self.logger.debug(f"📋 Encabezado fallback columna {column}: '{header_text}'")
                            return header_text
            
            self.logger.warning(f"❌ No se encontró encabezado válido para columna {column}")
            return None
            
        except Exception as e:
            self.logger.error(f"Error obteniendo encabezado de columna {column}: {e}")
            return None

    def _get_extracted_fasecolda_codes(self) -> tuple:
        """
        Obtiene los códigos FASECOLDA extraídos automáticamente.
        
        Returns:
            tuple: (cf_code, ch_code) o (None, None) si no están disponibles
        """
        try:
            # Método 1: Acceder directamente al extractor global
            from ..shared.fasecolda_extractor import _global_extractor
            
            self.logger.debug(f"🔍 Verificando extractor global: {_global_extractor}")
            
            if _global_extractor:
                self.logger.debug(f"🔍 Códigos en extractor: {_global_extractor.codes}")
                
                if _global_extractor.codes:
                    codes = _global_extractor.codes
                    cf_code = codes.get('cf_code')
                    ch_code = codes.get('ch_code', '')
                    
                    if cf_code:
                        self.logger.info(f"✅ Códigos FASECOLDA extraídos obtenidos: CF={cf_code}, CH={ch_code}")
                        return cf_code, ch_code
            
            # Método 2: Verificar si hay códigos en variables de entorno
            import os
            env_cf = os.environ.get('EXTRACTED_CF_CODE')
            env_ch = os.environ.get('EXTRACTED_CH_CODE')
            
            if env_cf:
                self.logger.info(f"✅ Códigos FASECOLDA desde variables de entorno: CF={env_cf}, CH={env_ch or ''}")
                return env_cf, env_ch or ''
            
            # Método 3: Buscar en logs recientes para códigos reportados
            cf_code, ch_code = self._extract_codes_from_logs()
            if cf_code:
                self.logger.info(f"✅ Códigos FASECOLDA extraídos de logs: CF={cf_code}, CH={ch_code}")
                return cf_code, ch_code
            
            self.logger.warning("⚠️ No hay códigos FASECOLDA extraídos disponibles por ningún método")
            return None, None
                
        except Exception as e:
            self.logger.warning(f"⚠️ Error obteniendo códigos FASECOLDA extraídos: {e}")
            return None, None
    
    def _extract_codes_from_logs(self) -> tuple:
        """
        Intenta extraer códigos FASECOLDA de logs recientes como último recurso.
        
        Returns:
            tuple: (cf_code, ch_code) o (None, None) si no encuentra
        """
        try:
            import re
            import glob
            from pathlib import Path
            
            # Buscar archivos de log recientes del fasecolda_extractor
            logs_dir = self.base_path / "LOGS" / "fasecolda_extractor"
            if not logs_dir.exists():
                return None, None
            
            # Buscar el archivo de log más reciente
            log_files = list(logs_dir.glob("*.log"))
            if not log_files:
                return None, None
            
            # Ordenar por fecha de modificación (más reciente primero)
            log_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            
            # Leer el archivo más reciente
            latest_log = log_files[0]
            with open(latest_log, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Buscar patrón: "CF: 05636039 - CH: 05606132"
            pattern = r'CF:\s*(\d+).*?CH:\s*(\d+)'
            match = re.search(pattern, content)
            
            if match:
                cf_code = match.group(1)
                ch_code = match.group(2)
                self.logger.debug(f"🔍 Códigos encontrados en logs: CF={cf_code}, CH={ch_code}")
                return cf_code, ch_code
            
            return None, None
            
        except Exception as e:
            self.logger.debug(f"🔍 Error extrayendo códigos de logs: {e}")
            return None, None

    def _setup_vigencia_dates_and_coverage(self, worksheet):
        """
        Configura automáticamente las fechas de vigencia y calcula los días de cobertura.
        
        - Fecha inicio vigencia: Automáticamente la fecha de hoy
        - Días de cobertura: Calcula automáticamente (fecha fin - fecha inicio)
        - Prima anual IVA incluido: Usa los días calculados para la fórmula
        """
        try:
            self.logger.info("📅 Configurando fechas de vigencia y calculando días de cobertura...")
            
            # 1. Obtener todas las columnas de aseguradoras
            all_company_columns = []
            companies = ['sura', 'allianz', 'bolivar', 'solidaria', 'sbs']
            for company in companies:
                company_cols = self._find_company_columns(worksheet, company)
                all_company_columns.extend(company_cols)
            
            if not all_company_columns:
                self.logger.warning("❌ No se encontraron columnas de aseguradoras")
                return
            
            self.logger.info(f"✅ Columnas de aseguradoras encontradas: {sorted(set(all_company_columns))}")
            
            # 2. Configurar fecha inicio vigencia con la fecha de hoy
            fecha_inicio_row = self._find_cell_with_text_in_any_column(worksheet, 'fecha inicio vigencia', 'inicio vigencia')
            if fecha_inicio_row:
                fecha_hoy = datetime.now().strftime('%d/%m/%Y')
                for col in sorted(set(all_company_columns)):
                    self._write_to_cell_safe(worksheet, fecha_inicio_row, col, fecha_hoy)
                self.logger.info(f"✅ Fecha inicio vigencia: {fecha_hoy} en todas las columnas")
            else:
                self.logger.warning("❌ No se encontró la fila 'Fecha inicio vigencia'")
                return
            
            # 3. Buscar fecha fin vigencia para calcular días de cobertura
            fecha_fin_row = self._find_cell_with_text_in_any_column(worksheet, 'fecha fin vigencia', 'fecha fin de vigencia', 'fin vigencia')
            if not fecha_fin_row:
                self.logger.warning("❌ No se encontró la fila 'Fecha fin vigencia'")
                return
            
            # 4. Obtener fecha fin vigencia (buscar en varias columnas)
            fecha_fin_value = None
            for col in range(2, 10):  # Buscar en columnas B a I
                cell_value = worksheet.cell(row=fecha_fin_row, column=col).value
                if cell_value and str(cell_value).strip():
                    fecha_fin_value = cell_value
                    break
            
            if not fecha_fin_value:
                self.logger.warning("❌ No hay fecha fin vigencia configurada en la plantilla")
                return
            
            # 5. Calcular días de cobertura
            dias_cobertura = self._calculate_coverage_days(fecha_hoy, str(fecha_fin_value))
            if dias_cobertura <= 0:
                self.logger.warning(f"❌ Días de cobertura inválidos: {dias_cobertura}")
                return
            
            # 6. Llenar días de cobertura en todas las columnas
            dias_cobertura_row = self._find_cell_with_text_in_any_column(worksheet, 'días de cobertura', 'dias de cobertura')
            if dias_cobertura_row:
                for col in sorted(set(all_company_columns)):
                    self._write_to_cell_safe(worksheet, dias_cobertura_row, col, dias_cobertura)
                self.logger.info(f"✅ Días de cobertura: {dias_cobertura} días en todas las columnas")
            else:
                self.logger.warning("❌ No se encontró la fila 'Días de cobertura'")
            
            # 7. Calcular prima anual en todas las columnas
            self._calculate_prima_anual_all_columns(worksheet, dias_cobertura, all_company_columns)
            
        except Exception as e:
            self.logger.error(f"Error configurando fechas y días de cobertura: {e}")

    def _find_target_column_for_row(self, worksheet, row: int) -> int:
        """
        Encuentra la columna objetivo para llenar datos en una fila específica.
        Busca la primera columna vacía después de la columna A.
        """
        # Empezar desde columna B (2) y buscar la primera vacía o la segunda columna
        for col in range(2, 6):  # Columnas B, C, D, E
            cell_value = worksheet.cell(row=row, column=col).value
            if not cell_value or str(cell_value).strip() == "":
                return col
        
        # Si todas están ocupadas, usar columna B (2) por defecto
        return 2

    def _calculate_prima_anual_all_columns(self, worksheet, dias_cobertura: int, company_columns: List[int]):
        """
        Calcula la prima anual en todas las columnas de aseguradoras.
        """
        try:
            self.logger.info(f"📊 Calculando prima anual en todas las columnas con {dias_cobertura} días de cobertura...")
            
            # 1. Buscar las filas necesarias
            valor_pagar_row = self._find_cell_with_text_in_any_column(worksheet, 'valor a pagar')
            prima_anual_row = self._find_cell_with_text_in_any_column(worksheet, 'prima anual')
            
            if not valor_pagar_row or not prima_anual_row:
                self.logger.warning("❌ No se encontraron las filas necesarias para calcular prima anual")
                return
            
            # 2. Calcular prima anual para cada columna de aseguradora
            valores_calculados = 0
            for col in sorted(set(company_columns)):
                try:
                    # Obtener valor a pagar de esta columna
                    valor_pagar_cell = worksheet.cell(row=valor_pagar_row, column=col)
                    valor_pagar = valor_pagar_cell.value
                    
                    if valor_pagar and str(valor_pagar).strip() and str(valor_pagar).strip() != "":
                        # Extraer valor numérico
                        valor_numerico = self._extract_numeric_value(str(valor_pagar))
                        if valor_numerico > 0:
                            # Calcular prima anual: (Valor a pagar) / (Días de cobertura) * 365
                            prima_anual_calculada = (valor_numerico / dias_cobertura) * 365
                            prima_formateada = f"${prima_anual_calculada:,.0f}".replace(",", ".")
                            
                            # Escribir en la misma columna
                            self._write_to_cell_safe(worksheet, prima_anual_row, col, prima_formateada)
                            valores_calculados += 1
                            
                            self.logger.debug(f"✅ Prima anual columna {col}: {prima_formateada}")
                
                except Exception as e:
                    self.logger.warning(f"⚠️ Error calculando prima anual en columna {col}: {e}")
                    continue
            
            if valores_calculados > 0:
                self.logger.info(f"✅ Prima anual calculada en {valores_calculados} columnas")
            else:
                self.logger.warning("⚠️ No se calcularon primas anuales")
                
        except Exception as e:
            self.logger.error(f"Error calculando prima anual en todas las columnas: {e}")

    def _calculate_prima_anual(self, worksheet, dias_cobertura: int):
        """
        Calcula la prima anual usando el valor existente de "VALOR A PAGAR (IVA INCLUIDO)".
        """
        try:
            self.logger.info(f"📊 Calculando prima anual con {dias_cobertura} días de cobertura...")
            
            # 1. Buscar la fila "VALOR A PAGAR (IVA INCLUIDO)"
            valor_pagar_row = self._find_cell_with_text_in_any_column(worksheet, 'valor a pagar')
            if not valor_pagar_row:
                self.logger.warning("❌ No se encontró la fila 'VALOR A PAGAR (IVA INCLUIDO)'")
                return
            
            # 2. Buscar la fila "PRIMA ANUAL IVA INCLUIDO"
            prima_anual_row = self._find_cell_with_text_in_any_column(worksheet, 'prima anual')
            if not prima_anual_row:
                self.logger.warning("❌ No se encontró la fila 'PRIMA ANUAL IVA INCLUIDO'")
                return
            
            # 3. Obtener el valor a pagar (buscar en varias columnas)
            valor_pagar = None
            valor_col = None
            for col in range(2, 10):  # Buscar en columnas B a I
                cell_value = worksheet.cell(row=valor_pagar_row, column=col).value
                if cell_value and str(cell_value).strip() and str(cell_value).strip() != "":
                    valor_pagar = cell_value
                    valor_col = col
                    break
            
            if not valor_pagar:
                self.logger.warning("❌ No se encontró valor a pagar para calcular prima anual")
                return
            
            # 4. Calcular prima anual
            valor_numerico = self._extract_numeric_value(str(valor_pagar))
            if valor_numerico > 0:
                # Calcular: (Valor a pagar) / (Días de cobertura) * 365
                prima_anual_calculada = (valor_numerico / dias_cobertura) * 365
                
                # Formatear como moneda
                prima_formateada = f"${prima_anual_calculada:,.0f}".replace(",", ".")
                
                # Escribir en la misma columna donde está el valor a pagar
                self._write_to_cell_safe(worksheet, prima_anual_row, valor_col, prima_formateada)
                
                self.logger.info(f"✅ Prima anual calculada: ${valor_numerico:,.0f} ÷ {dias_cobertura} × 365 = {prima_formateada}")
            else:
                self.logger.warning(f"❌ Valor numérico inválido extraído: {valor_numerico}")
                
        except Exception as e:
            self.logger.error(f"Error calculando prima anual: {e}")

    def _calculate_coverage_days(self, fecha_inicio: str, fecha_fin: str) -> int:
        """
        Calcula los días de cobertura entre dos fechas.
        
        Args:
            fecha_inicio: Fecha de inicio en formato DD/MM/YYYY
            fecha_fin: Fecha de fin en formato DD/MM/YYYY o datetime
            
        Returns:
            int: Número de días de cobertura
        """
        try:
            # Parsear fecha inicio
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%d/%m/%Y')
            
            # Parsear fecha fin (puede venir en diferentes formatos)
            if isinstance(fecha_fin, datetime):
                fecha_fin_obj = fecha_fin
            else:
                # Intentar diferentes formatos de fecha
                fecha_fin_str = str(fecha_fin).strip()
                try:
                    fecha_fin_obj = datetime.strptime(fecha_fin_str, '%d/%m/%Y')
                except ValueError:
                    try:
                        fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
                    except ValueError:
                        try:
                            # Formato con hora
                            fecha_fin_obj = datetime.strptime(fecha_fin_str.split(' ')[0], '%Y-%m-%d')
                        except ValueError:
                            self.logger.error(f"No se pudo parsear fecha fin: {fecha_fin_str}")
                            return 0
            
            # Calcular diferencia en días
            diferencia = fecha_fin_obj - fecha_inicio_obj
            dias = diferencia.days
            
            # Asegurar que sea positivo
            if dias <= 0:
                self.logger.warning(f"Días negativos o cero calculados: {dias}")
                return 0
            
            self.logger.info(f"📅 Cálculo: {fecha_inicio} hasta {fecha_fin_obj.strftime('%d/%m/%Y')} = {dias} días")
            return dias
            
        except Exception as e:
            self.logger.error(f"Error calculando días de cobertura: {e}")
            return 0

    def _recalculate_annualized_values(self, worksheet, dias_cobertura: int):
        """
        Recalcula los valores anualizados usando los días de cobertura calculados automáticamente.
        """
        try:
            self.logger.info(f"📊 Recalculando valores anualizados con {dias_cobertura} días de cobertura...")
            
            # 1. Buscar la fila "VALOR A PAGAR (IVA INCLUIDO)"
            valor_pagar_row = self._find_cell_with_text_in_any_column(worksheet, 'valor a pagar', 'iva incluido')
            if not valor_pagar_row:
                self.logger.warning("❌ No se encontró la fila 'VALOR A PAGAR (IVA INCLUIDO)'")
                return
            
            # 2. Buscar la fila "PRIMA ANUAL IVA INCLUIDO"
            prima_anual_row = self._find_cell_with_text_in_any_column(worksheet, 'prima anual', 'iva incluido')
            if not prima_anual_row:
                self.logger.warning("❌ No se encontró la fila 'PRIMA ANUAL IVA INCLUIDO'")
                return
            
            # 3. Recorrer las columnas y calcular valores anualizados
            valores_calculados = 0
            for col in range(2, 20):  # Columnas B hasta S
                try:
                    # Obtener valor a pagar de la fila correspondiente
                    valor_pagar_cell = worksheet.cell(row=valor_pagar_row, column=col)
                    valor_pagar = valor_pagar_cell.value
                    
                    # Solo calcular si el valor existe y es válido
                    if valor_pagar and str(valor_pagar).strip() != "":
                        # Limpiar valor a pagar (quitar $ y puntos)
                        valor_numerico = self._extract_numeric_value(str(valor_pagar))
                        
                        if valor_numerico > 0:
                            # Calcular valor anualizado: (Valor a pagar) / (Días de cobertura) * 365
                            valor_anualizado = (valor_numerico / dias_cobertura) * 365
                            
                            # Formatear como moneda
                            valor_formateado = f"${valor_anualizado:,.0f}".replace(",", ".")
                            
                            # Colocar en la fila PRIMA ANUAL
                            self._write_to_cell_safe(worksheet, prima_anual_row, col, valor_formateado)
                            
                            valores_calculados += 1
                            self.logger.info(f"✅ Columna {col}: ${valor_numerico:,.0f} ÷ {dias_cobertura} × 365 = {valor_formateado}")
                
                except Exception as e:
                    self.logger.warning(f"⚠️ Error calculando valor anualizado en columna {col}: {e}")
                    continue
            
            if valores_calculados > 0:
                self.logger.info(f"✅ Total de valores anualizados calculados: {valores_calculados}")
            else:
                self.logger.warning("⚠️ No se calcularon valores anualizados")
                
        except Exception as e:
            self.logger.error(f"Error recalculando valores anualizados: {e}")
